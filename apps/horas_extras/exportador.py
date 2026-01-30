# apps/horas_extras/exportador.py
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse

# Traducciones al español
DIAS_ES = {
    'Monday': 'Lunes',
    'Tuesday': 'Martes',
    'Wednesday': 'Miércoles',
    'Thursday': 'Jueves',
    'Friday': 'Viernes',
    'Saturday': 'Sábado',
    'Sunday': 'Domingo',
}

MESES_ES = {
    'January': 'Enero',
    'February': 'Febrero',
    'March': 'Marzo',
    'April': 'Abril',
    'May': 'Mayo',
    'June': 'Junio',
    'July': 'Julio',
    'August': 'Agosto',
    'September': 'Septiembre',
    'October': 'Octubre',
    'November': 'Noviembre',
    'December': 'Diciembre',
}

def traducir_dia(dia_en):
    """Traduce un nombre de día de inglés a español."""
    return DIAS_ES.get(dia_en, dia_en)

def traducir_periodo(periodo_str):
    """Traduce el periodo (ej: 'January 2026' -> 'Enero 2026')."""
    for mes_en, mes_es in MESES_ES.items():
        periodo_str = periodo_str.replace(mes_en, mes_es)
    return periodo_str

class ExportadorReportes:
    """Exportador de reportes de horas extras y recargos a Excel (Norma Colombiana)"""

    @classmethod
    def generar_excel(cls, datos_reporte, totales_generales, periodo_str):
        """
        Genera un archivo Excel con el reporte unificado (HOD, RNO, RDF, RNF).
        
        Args:
            datos_reporte: Lista de diccionarios con info de operadores y sus días.
            totales_generales: Diccionario con sumatorias globales.
            periodo_str: Texto del periodo (ej: "Enero 2026").
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Reporte Horas Extras'
        
        # Estilos
        font_header = Font(bold=True, size=12, color="FFFFFF")
        fill_header = PatternFill(start_color="0056b3", end_color="0056b3", fill_type="solid")
        font_subtotal = Font(bold=True)
        fill_subtotal = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Título Principal
        ws.merge_cells('A1:H1')
        periodo_es = traducir_periodo(periodo_str)
        ws['A1'] = f'Reporte de Horas Extras y Recargos - {periodo_es}'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        current_row = 3
        
        # Iterar por cada operador
        for operador in datos_reporte:
            # Encabezado Operador
            ws.merge_cells(f'A{current_row}:H{current_row}')
            ws[f'A{current_row}'] = f"Operador: {operador['nombre']} (ID: {operador['id']})"
            ws[f'A{current_row}'].font = Font(bold=True, size=11)
            ws[f'A{current_row}'].fill = PatternFill(start_color="cce5ff", end_color="cce5ff", fill_type="solid")
            current_row += 1
            
            # Encabezados de Tabla
            headers = ['Fecha', 'Día', 'Turno', 'HOD', 'RNO', 'RDF', 'RNF', 'Total']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = header
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal='center')
                cell.border = border_thin
            current_row += 1
            
            # Datos de Días
            for dia in operador['dias']:
                ws.cell(row=current_row, column=1, value=dia['fecha']) # Fecha
                ws.cell(row=current_row, column=2, value=traducir_dia(dia['dia_semana'])) # Día en español
                ws.cell(row=current_row, column=3, value=dia['turno']) # Turno
                
                # Horas (numérico)
                ws.cell(row=current_row, column=4, value=float(dia['horas']['HOD']))
                ws.cell(row=current_row, column=5, value=float(dia['horas']['RNO']))
                ws.cell(row=current_row, column=6, value=float(dia['horas']['RDF']))
                ws.cell(row=current_row, column=7, value=float(dia['horas']['RNF']))
                ws.cell(row=current_row, column=8, value=float(dia['horas']['TOTAL']))
                
                # Estilo fila
                for col in range(1, 9):
                    ws.cell(row=current_row, column=col).border = border_thin
                    if col >= 4: # Horas centradas
                        ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='center')
                
                # Resaltar festivos
                if dia['es_festivo']:
                    for col in range(1, 9):
                         ws.cell(row=current_row, column=col).font = Font(color="dc3545") # Rojo
                
                current_row += 1
            
            # Subtotales Operador
            ws.cell(row=current_row, column=3, value="TOTALES").font = font_subtotal
            ws.cell(row=current_row, column=3).alignment = Alignment(horizontal='right')
            
            totales = operador['totales']
            ws.cell(row=current_row, column=4, value=float(totales['HOD'])).font = font_subtotal
            ws.cell(row=current_row, column=5, value=float(totales['RNO'])).font = font_subtotal
            ws.cell(row=current_row, column=6, value=float(totales['RDF'])).font = font_subtotal
            ws.cell(row=current_row, column=7, value=float(totales['RNF'])).font = font_subtotal
            ws.cell(row=current_row, column=8, value=float(totales['TOTAL'])).font = font_subtotal
            
            for col in range(1, 9):
                ws.cell(row=current_row, column=col).fill = fill_subtotal
                ws.cell(row=current_row, column=col).border = border_thin
            
            current_row += 3 # Espacio entre operadores
            
        # Ajuste de columnas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 10
        
        # Preparar respuesta
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        clean_periodo = periodo_str.replace(" ", "_")
        response['Content-Disposition'] = f'attachment; filename="Reporte_Horas_{clean_periodo}.xlsx"'
        wb.save(response)
        return response
