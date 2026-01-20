# apps/reportes/views.py
import csv
from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator
from django.utils import timezone
from django.urls import reverse
from django.contrib import messages
from datetime import datetime, timedelta
import json
import logging

# Para exportación Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# =============================================================================
# FUNCIÓN AUXILIAR PARA CREAR EXCEL CON FORMATO PROFESIONAL
# =============================================================================

def create_styled_excel(data, headers, filename, title=None, sheet_name='Datos'):
    """
    Crea un archivo Excel con formato profesional.
    
    Args:
        data: Lista de diccionarios con los datos
        headers: Lista de tuplas (key, label)
        filename: Nombre del archivo
        title: Título del reporte (opcional)
        sheet_name: Nombre de la hoja
    
    Returns:
        HttpResponse con el archivo Excel
    """
    if not OPENPYXL_AVAILABLE:
        # Fallback a CSV si openpyxl no está disponible
        return create_csv_response(data, headers, filename)
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Estilos
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    even_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    odd_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    header_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    start_row = 1
    
    # Agregar título si existe
    if title:
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(bold=True, size=14, color='1F4E79')
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        date_cell = ws.cell(row=2, column=1, value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        date_cell.font = Font(italic=True, size=10, color='666666')
        
        start_row = 4
    
    # Escribir encabezados
    header_keys = [h[0] for h in headers]
    header_labels = [h[1] for h in headers]
    
    for col_idx, label in enumerate(header_labels, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border
    
    # Escribir datos
    for row_idx, row_data in enumerate(data, start_row + 1):
        for col_idx, key in enumerate(header_keys, 1):
            value = row_data.get(key, '') if isinstance(row_data, dict) else ''
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Alternar colores de fila
            if (row_idx - start_row) % 2 == 0:
                cell.fill = even_fill
            else:
                cell.fill = odd_fill
            
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Ajustar ancho de columnas
    for col_idx, label in enumerate(header_labels, 1):
        column_letter = get_column_letter(col_idx)
        max_length = len(str(label))
        
        for row in ws.iter_rows(min_row=start_row + 1, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
        
        ws.column_dimensions[column_letter].width = min(max_length + 3, 50)
    
    # Congelar encabezados
    ws.freeze_panes = f'A{start_row + 1}'
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


def create_csv_response(data, headers, filename):
    """Crea una respuesta HTTP con un archivo CSV formateado correctamente."""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # BOM para que Excel reconozca UTF-8
    response.write('\ufeff')
    
    # Usar punto y coma como delimitador para mejor compatibilidad con Excel
    writer = csv.writer(response, delimiter=';', quoting=csv.QUOTE_MINIMAL)
    
    # Escribir encabezados
    header_labels = [h[1] for h in headers]
    writer.writerow(header_labels)
    
    # Escribir datos
    header_keys = [h[0] for h in headers]
    for row_data in data:
        if isinstance(row_data, dict):
            row = [row_data.get(key, '') for key in header_keys]
        else:
            row = row_data
        writer.writerow(row)
    
    return response

from .utils import (
    ejecutar_procedimiento_almacenado,
    ejecutar_consulta_personalizada,
    ejecutar_sp_dashboard_metrics,
    obtener_servidores_disponibles,
    obtener_bases_datos,
    formatear_resultado_backup,
    calcular_estadisticas_cumplimiento
)
from .config import STORED_PROCEDURES, QUERIES, DEFAULT_FILTERS, PAGINATION, THRESHOLDS
from .data_converters import (
    convert_cumplimiento_result,
    convert_jobs_result,
    normalize_results,
    format_cumplimiento_results
)
from .constants import (
    DateFormats,
    ExportHeaders,
    ExportFileNames,
    SheetNames,
    ReportTitles
)

logger = logging.getLogger(__name__)




@login_required
def buscar_reporte_cumplimiento(request):
    """Buscar resultados de cumplimiento con fechas específicas"""
    try:
        fecha_inicio_param = request.GET.get('fecha')
        fecha_fin_param = request.GET.get('fecha1')
        
        if not fecha_inicio_param or not fecha_fin_param:
            # Redirigir de vuelta con mensaje de error
            messages.error(request, 'Debe proporcionar ambas fechas para realizar la búsqueda.')
            return redirect('reportes:cumplimiento_backup')
        
        # Redirigir a la vista principal con los parámetros
        return redirect(f"{reverse('reportes:cumplimiento_backup')}?fecha={fecha_inicio_param}&fecha1={fecha_fin_param}")
        
    except Exception as e:
        logger.error(f"Error en búsqueda: {e}")
        messages.error(request, f'Error en la búsqueda: {str(e)}')
        return redirect('reportes:cumplimiento_backup')


@login_required
def reporte_cumplimiento_excel(request):
    """Generar reporte de cumplimiento en Excel para descarga"""
    try:
        from io import StringIO
        
        fecha_inicio_param = request.GET.get('fecha')
        fecha_fin_param = request.GET.get('fecha1')
        
        # Configurar fechas
        if fecha_inicio_param:
            fecha_inicio = datetime.strptime(fecha_inicio_param, '%Y-%m-%d').strftime('%Y-%m-%d')
        else:
            fecha_inicio = datetime.now().replace(day=1).strftime('%Y-%m-%d')
            
        if fecha_fin_param:
            fecha_fin = datetime.strptime(fecha_fin_param, '%Y-%m-%d').strftime('%Y-%m-%d')
        else:
            fecha_fin = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
        
        # Obtener datos usando el mismo procedimiento
        try:
            resultados = ejecutar_procedimiento_almacenado(
                'sp_Programaciondebcks',
                [fecha_inicio, fecha_fin]
            )

            # Convertir a formato consistente si es necesario
            resultados = normalize_results(resultados, convert_cumplimiento_result)
                
        except Exception as sp_error:
            logger.error(f"Error en SP para Excel: {sp_error}")
            # Usar consulta alternativa desde config
            resultados = ejecutar_consulta_personalizada(QUERIES['cumplimiento_fallback'], [fecha_inicio, fecha_fin])
        
        # Crear respuesta CSV
        response = HttpResponse(content_type='text/csv')
        filename = f"cumplimiento_backup_{fecha_inicio.replace('/', '-')}_a_{fecha_fin.replace('/', '-')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Escribir CSV
        writer = csv.writer(response)
        writer.writerow([
            'Servidor',
            'Nombre BD', 
            'IP Servidor',
            'Copias Ejecutadas',
            'Copias Programadas',
            'Porcentaje Cumplimiento'
        ])
        
        for resultado in resultados:
            writer.writerow([
                resultado.get('SERVIDOR', ''),
                resultado.get('DatabaseName', ''),
                resultado.get('IPSERVER', ''),
                resultado.get('TOTAL', ''),
                resultado.get('TOTALPROGRAM', ''),
                resultado.get('PORCENTAJE', '')
            ])
        
        logger.info(f"Descarga de reporte: {filename}")
        return response
        
    except Exception as e:
        logger.error(f"Error generando Excel: {e}")
        messages.error(request, f'Error al generar el reporte: {str(e)}')
        return redirect('reportes:cumplimiento_backup')


@login_required
def dashboard_view(request):
    """Dashboard principal usando sp_DashboardMetrics"""
    try:
        # Ejecutar el SP que devuelve todos los datos del dashboard
        datos_dashboard = ejecutar_sp_dashboard_metrics()
        
        # Extraer métricas principales
        metricas = datos_dashboard.get('metricas', {})
        
        # Preparar datos para los gráficos
        # Gráfico de Jobs (dona)
        jobs_stats = []
        for job in datos_dashboard.get('stats_jobs', []):
            jobs_stats.append({
                'label': job.get('resultado_agrupado', 'Otro'),
                'value': job.get('cantidad', 0)
            })
        
        # Gráfico de Tipos de Backup (polar)
        backup_types = []
        for tipo in datos_dashboard.get('tipos_backup', []):
            backup_types.append({
                'label': tipo.get('tipo_backup', 'Sin Tipo'),
                'value': tipo.get('cantidad', 0)
            })
        
        # Gráfico de Tendencia Semanal (línea)
        weekly_trend = []
        for dia in datos_dashboard.get('tendencia_semanal', []):
            weekly_trend.append({
                'day': dia.get('dia_semana', ''),
                'backups': dia.get('cantidad_backups', 0)
            })
        
        # Gráfico de Top Servidores (barras)
        top_servers = []
        for servidor in datos_dashboard.get('top_servidores', []):
            top_servers.append({
                'server': servidor.get('SERVIDOR', 'N/A'),
                'backups': servidor.get('total_backups', 0)
            })
        
        # Obtener últimos backups desde sp_ultimosbck para la lista lateral
        try:
            ultimos_backups = ejecutar_procedimiento_almacenado('sp_ultimosbck')
            # Limitar a 10 registros y agregar status_class
            ultimos_backups = ultimos_backups[:10] if ultimos_backups else []
            for backup in ultimos_backups:
                # Agregar clase de estado basada en la fecha (todos serán 'danger' por ser históricos)
                backup['status_class'] = 'danger'
        except Exception as e:
            logger.warning(f"Error obteniendo últimos backups: {e}")
            ultimos_backups = []

        context = {
            'metricas': metricas,
            'stats_jobs': datos_dashboard.get('stats_jobs', []),
            'ultimos_backups': ultimos_backups,
            'chart_data': {
                'servidores': metricas.get('total_servidores', 0),
                'bases_datos': metricas.get('total_bases_datos', 0),
                'backups_hoy': metricas.get('backups_hoy', 0),
                'backups_semana': metricas.get('backups_semana', 0),
                # Datos para gráficos en formato JSON
                'jobs_stats': json.dumps(jobs_stats) if jobs_stats else '[]',
                'backup_types': json.dumps(backup_types) if backup_types else '[]',
                'weekly_trend': json.dumps(weekly_trend) if weekly_trend else '[]',
                'top_servers': json.dumps(top_servers) if top_servers else '[]',
            },
            'last_updated': timezone.now(),
        }

        return render(request, 'reportes/dashboard.html', context)

    except Exception as e:
        logger.error(f"Error en dashboard: {e}")
        context = {
            'error': f'Error al cargar las métricas del dashboard: {str(e)}',
            'metricas': {},
            'stats_jobs': [],
            'ultimos_backups': [],
            'chart_data': {
                'servidores': 0, 
                'bases_datos': 0, 
                'backups_hoy': 0, 
                'backups_semana': 0,
                'jobs_stats': '[]',
                'backup_types': '[]',
                'weekly_trend': '[]',
                'top_servers': '[]',
            }
        }
        return render(request, 'reportes/dashboard.html', context)


@login_required
def cumplimiento_backup_view(request):
    """Reporte de cumplimiento usando sp_Programaciondebcks con fechas personalizables"""
    try:
        # Obtener fechas de los parámetros GET
        fecha_inicio_param = request.GET.get('fecha')
        fecha_fin_param = request.GET.get('fecha1')
        
        # Configurar fechas por defecto si no se proporcionan
        if fecha_inicio_param:
            try:
                fecha_inicio = datetime.strptime(fecha_inicio_param, '%Y-%m-%d').strftime('%Y-%m-%d')
            except ValueError:
                logger.warning(f"Fecha inicio inválida: {fecha_inicio_param}")
                fecha_inicio = datetime.now().replace(day=1).strftime('%Y-%m-%d')
        else:
            # Primer día del mes actual
            fecha_inicio = datetime.now().replace(day=1).strftime('%Y-%m-%d')
        
        if fecha_fin_param:
            try:
                fecha_fin = datetime.strptime(fecha_fin_param, '%Y-%m-%d').strftime('%Y-%m-%d')
            except ValueError:
                logger.warning(f"Fecha fin inválida: {fecha_fin_param}")
                fecha_fin = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
        else:
            # Día anterior
            fecha_fin = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
        
        logger.info(f"Ejecutando cumplimiento backup: {fecha_inicio} a {fecha_fin}")
        
        # Ejecutar el procedimiento almacenado
        try:
            resultados = ejecutar_procedimiento_almacenado(
                'sp_Programaciondebcks',
                [fecha_inicio, fecha_fin]
            )

            # Normalizar resultados usando data_converters
            resultados = normalize_results(resultados, convert_cumplimiento_result)

        except Exception as sp_error:
            logger.error(f"Error ejecutando sp_Programaciondebcks: {sp_error}")
            # Consulta alternativa desde config si el SP falla
            resultados = ejecutar_consulta_personalizada(
                QUERIES['cumplimiento_fallback'],
                [fecha_inicio, fecha_fin]
            )
        
        # Estadísticas para el dashboard
        total_registros = len(resultados)
        total_ejecutadas = sum(int(r.get('TOTAL', 0)) for r in resultados)
        total_programadas = sum(int(r.get('TOTALPROGRAM', 0)) for r in resultados)
        
        estadisticas = {
            'total_registros': total_registros,
            'total_ejecutadas': total_ejecutadas,
            'total_programadas': total_programadas,
            'promedio_cumplimiento': round((total_ejecutadas / total_programadas) * 100, 2) if total_programadas > 0 else 0,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin
        }
        
        context = {
            'resultadosCump': resultados,
            'estadisticas': estadisticas,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'total_registros': total_registros
        }
        
        return render(request, 'reportes/cumplimiento_bck.html', context)
        
    except Exception as e:
        logger.error(f"Error en cumplimiento backup: {e}")
        context = {
            'error': f'Error al cargar el reporte de cumplimiento: {str(e)}',
            'resultadosCump': [],
            'estadisticas': {
                'total_registros': 0,
                'total_ejecutadas': 0, 
                'total_programadas': 0,
                'promedio_cumplimiento': 0
            }
        }
        return render(request, 'reportes/cumplimiento_bck.html', context)


@login_required
def jobs_backup_view(request):
    """Reporte de jobs usando sp_resultadoJobsBck - coincide exactamente con los campos del SP"""
    try:
        # Filtros
        fecha_inicio = request.GET.get('fecha_inicio', (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
        fecha_fin = request.GET.get('fecha_fin', timezone.now().strftime('%Y-%m-%d'))
        servidor = request.GET.get('servidor', '')
        resultado_filtro = request.GET.get('resultado', '')

        logger.info(f"Ejecutando jobs backup: {fecha_inicio} a {fecha_fin}")

        # Usar procedimiento almacenado sp_resultadoJobsBck
        try:
            resultados = ejecutar_procedimiento_almacenado(
                'sp_resultadoJobsBck',
                [fecha_inicio, fecha_fin]
            )

            # Normalizar resultados usando data_converters
            resultados = normalize_results(resultados, convert_jobs_result)

        except Exception as proc_error:
            logger.warning(f"Error con sp_resultadoJobsBck: {proc_error}")
            # Consulta directa como alternativa desde config
            resultados = ejecutar_consulta_personalizada(QUERIES['jobs_resultado_directo'], [fecha_inicio, fecha_fin])

        # Aplicar filtros adicionales en Python si es necesario
        if servidor:
            resultados = [r for r in resultados if servidor.lower() in r.get('SERVIDOR', '').lower()]
        if resultado_filtro:
            resultados = [r for r in resultados if resultado_filtro.lower() in r.get('RESULTADO', '').lower()]

        # Estadísticas
        total = len(resultados)
        exitosos = len([r for r in resultados if 'exitoso' in r.get('RESULTADO', '').lower()])
        fallidos = len([r for r in resultados if 'fallido' in r.get('RESULTADO', '').lower() or 'error' in r.get('RESULTADO', '').lower()])
        otros = total - exitosos - fallidos
        
        stats = {
            'total': total,
            'exitosos': exitosos,
            'fallidos': fallidos,
            'otros': otros,
            'porcentaje_exito': round((exitosos / total) * 100, 2) if total > 0 else 0
        }

        # Paginación
        paginator = Paginator(resultados, PAGINATION['items_per_page'])
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Obtener lista de servidores para el filtro
        try:
            servidores = ejecutar_consulta_personalizada(QUERIES['servidores_jobs'])
        except Exception:
            servidores = []

        # Obtener tipos de resultado para referencia
        try:
            tipos_resultado = ejecutar_consulta_personalizada(QUERIES['tipos_resultado_jobs'])
        except Exception:
            tipos_resultado = []

        context = {
            'resultados': page_obj,
            'stats': stats,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'servidor': servidor,
            'resultado': resultado_filtro,
            'servidores': servidores,
            'tipos_resultado': tipos_resultado,
        }

        return render(request, 'reportes/jobs_bck.html', context)

    except Exception as e:
        logger.error(f"Error en jobs backup: {e}")
        context = {
            'error': f'Error al cargar los jobs de backup: {str(e)}',
            'resultados': [],
            'stats': {'total': 0, 'exitosos': 0, 'fallidos': 0, 'otros': 0, 'porcentaje_exito': 0},
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'servidor': '',
            'resultado': '',
            'servidores': [],
            'tipos_resultado': []
        }
        return render(request, 'reportes/jobs_bck.html', context)


@login_required
def archivos_backup_view(request):
    """Reporte de archivos usando BACKUPSGENERADOS"""
    try:
        # Filtros
        dias_atras = int(request.GET.get('dias_atras', 30))
        servidor = request.GET.get('servidor', '')
        tipo_backup = request.GET.get('tipo_backup', '')

        query = """
            SELECT 
                BCK_ID,
                SERVIDOR,
                DatabaseName as database_name,
                IPSERVER,
                TYPE as tipo_backup,
                FECHA,
                HORA,
                physical_device_name as ruta_archivo,
                CASE TYPE
                    WHEN 'FULL' THEN 'Completo'
                    WHEN 'INCREMENTAL' THEN 'Incremental'
                    WHEN 'DIFF' THEN 'Diferencial'
                    WHEN 'LOG' THEN 'Log'
                    ELSE ISNULL(TYPE, 'Sin tipo')
                END as tipo_descripcion,
                LEN(physical_device_name) as longitud_ruta
            FROM BACKUPSGENERADOS
            WHERE CONVERT(date, SUBSTRING(FECHA,7,4) + '-' + SUBSTRING(FECHA,4,2) + '-' + SUBSTRING(FECHA,1,2)) >= DATEADD(day, -%s, GETDATE())
                AND physical_device_name IS NOT NULL
        """
        
        params = [dias_atras]
        
        if servidor:
            query += " AND SERVIDOR = %s"
            params.append(servidor)
        if tipo_backup:
            query += " AND TYPE = %s"
            params.append(tipo_backup)
            
        query += """
            ORDER BY 
                CONVERT(date, SUBSTRING(FECHA,7,4) + '-' + SUBSTRING(FECHA,4,2) + '-' + SUBSTRING(FECHA,1,2)) DESC,
                CONVERT(time, HORA) DESC
        """
        
        resultados = ejecutar_consulta_personalizada(query, params)

        # Paginación
        paginator = Paginator(resultados, PAGINATION['items_per_page'])
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context = {
            'resultados': page_obj,
            'dias_atras': dias_atras,
            'servidor': servidor,
            'tipo_backup': tipo_backup,
            'servidores': ejecutar_consulta_personalizada("""
                SELECT DISTINCT SERVIDOR as servidor, IPSERVER as ip_servidor, COUNT(*) as total_backups
                FROM BACKUPSGENERADOS 
                WHERE SERVIDOR IS NOT NULL 
                GROUP BY SERVIDOR, IPSERVER 
                ORDER BY SERVIDOR
            """),
            'tipos_backup': ejecutar_consulta_personalizada("""
                SELECT DISTINCT TYPE as tipo_backup, COUNT(*) as cantidad
                FROM BACKUPSGENERADOS 
                WHERE TYPE IS NOT NULL
                GROUP BY TYPE 
                ORDER BY cantidad DESC
            """),
        }

        return render(request, 'reportes/archivos_bak.html', context)

    except Exception as e:
        logger.error(f"Error en archivos backup: {e}")
        context = {
            'error': f'Error al cargar los archivos de backup: {str(e)}',
            'resultados': []
        }
        return render(request, 'reportes/archivos_bak.html', context)


@login_required
def estados_db_view(request):
    """Reporte de estados usando sp_MonitorDatabaseStatus y DatabaseStatusLog"""
    try:
        # Filtros
        servidor = request.GET.get('servidor', '')
        estado = request.GET.get('estado', '')

        logger.info("Ejecutando estados de bases de datos con sp_MonitorDatabaseStatus")

        # Usar procedimiento almacenado sp_MonitorDatabaseStatus y consultar DatabaseStatusLog
        try:
            # Primero ejecutar el SP para actualizar datos
            try:
                ejecutar_procedimiento_almacenado('sp_MonitorDatabaseStatus')
                logger.info("SP sp_MonitorDatabaseStatus ejecutado exitosamente")
            except Exception as sp_error:
                # El SP puede no devolver resultados (solo hace INSERT), esto es normal
                logger.info(f"SP ejecutado (sin resultados de SELECT): {sp_error}")
            
            # Consultar datos desde config
            query = QUERIES['estados_db_log']
            params = []
            
            if servidor:
                query += " AND ServerName LIKE %s"
                params.append(f'%{servidor}%')
            if estado:
                query += " AND StateDesc LIKE %s"
                params.append(f'%{estado}%')
                
            query += " ORDER BY LastLogDate DESC, ServerName, DatabaseName"
            
            resultados = ejecutar_consulta_personalizada(query, params)
            
        except Exception as proc_error:
            logger.warning(f"Error con sp_MonitorDatabaseStatus: {proc_error}")
            # Consulta directa desde config como alternativa
            resultados = ejecutar_consulta_personalizada(QUERIES['estados_db_direct'])

        # Aplicar filtros adicionales en Python si es necesario
        if servidor:
            resultados = [r for r in resultados if servidor.lower() in r.get('SERVIDOR', '').lower()]
        if estado:
            resultados = [r for r in resultados if estado.lower() in r.get('ESTADO', '').lower()]

        # Estadísticas
        total = len(resultados)
        online = len([r for r in resultados if r.get('ESTADO', '').upper() == 'ONLINE'])
        servidores_unicos = len(set(r.get('SERVIDOR', '') for r in resultados if r.get('SERVIDOR')))
        otros = total - online
        
        stats = {
            'total': total,
            'online': online,
            'otros': otros,
            'servidores': servidores_unicos
        }

        # Paginación
        paginator = Paginator(resultados, PAGINATION['items_per_page'])
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Obtener lista de servidores para el filtro
        try:
            servidores = list(set(r.get('SERVIDOR', '') for r in resultados if r.get('SERVIDOR')))
            servidores = [{'servidor': srv} for srv in sorted(servidores)]
        except Exception:
            servidores = []

        # Obtener lista de estados para el filtro
        try:
            estados_unicos = list(set(r.get('ESTADO', '') for r in resultados if r.get('ESTADO')))
            estados = [{'estado': est} for est in sorted(estados_unicos)]
        except Exception:
            estados = []

        context = {
            'resultados': page_obj,
            'stats': stats,
            'servidor': servidor,
            'estado': estado,
            'servidores': servidores,
            'estados': estados,
        }

        return render(request, 'reportes/estados_db.html', context)

    except Exception as e:
        logger.error(f"Error en estados DB: {e}")
        context = {
            'error': f'Error al cargar los estados de bases de datos: {str(e)}',
            'resultados': [],
            'stats': {'total': 0, 'online': 0, 'otros': 0, 'servidores': 0},
            'servidor': '',
            'estado': '',
            'servidores': [],
            'estados': []
        }
        return render(request, 'reportes/estados_db.html', context)


@login_required
def ultimos_backup_view(request):
    """Reporte de últimos backups usando sp_ultimosbck"""
    try:
        # Usar procedimiento almacenado
        try:
            resultados = ejecutar_procedimiento_almacenado('sp_ultimosbck')
        except Exception as proc_error:
            logger.warning(f"Error con sp_ultimosbck: {proc_error}")
            # Consulta alternativa
            resultados = ejecutar_consulta_personalizada("""
                WITH UltimosBackupsPorBD AS (
                    SELECT 
                        SERVIDOR,
                        DatabaseName,
                        IPSERVER,
                        TYPE,
                        FECHA,
                        HORA,
                        ROW_NUMBER() OVER (
                            PARTITION BY SERVIDOR, DatabaseName 
                            ORDER BY 
                                CONVERT(date, SUBSTRING(FECHA,7,4) + '-' + SUBSTRING(FECHA,4,2) + '-' + SUBSTRING(FECHA,1,2)) DESC,
                                CONVERT(time, HORA) DESC
                        ) as rn,
                        DATEDIFF(hour, 
                            CONVERT(datetime, SUBSTRING(FECHA,7,4) + '-' + SUBSTRING(FECHA,4,2) + '-' + SUBSTRING(FECHA,1,2) + ' ' + HORA), 
                            GETDATE()
                        ) as horas_transcurridas
                    FROM BACKUPSGENERADOS
                    WHERE SERVIDOR IS NOT NULL AND DatabaseName IS NOT NULL
                )
                SELECT 
                    SERVIDOR,
                    DatabaseName,
                    IPSERVER,
                    TYPE,
                    FECHA,
                    HORA,
                    horas_transcurridas,
                    CASE 
                        WHEN horas_transcurridas <= 24 THEN 'success'
                        WHEN horas_transcurridas <= 48 THEN 'warning'
                        ELSE 'danger'
                    END as status_class
                FROM UltimosBackupsPorBD
                WHERE rn = 1
                ORDER BY 
                    CONVERT(date, SUBSTRING(FECHA,7,4) + '-' + SUBSTRING(FECHA,4,2) + '-' + SUBSTRING(FECHA,1,2)) DESC,
                    CONVERT(time, HORA) DESC
            """)

        context = {
            'resultados': resultados,
        }

        return render(request, 'reportes/ultimos_bck.html', context)

    except Exception as e:
        logger.error(f"Error en últimos backup: {e}")
        context = {
            'error': f'Error al cargar los últimos backups: {str(e)}',
            'resultados': []
        }
        return render(request, 'reportes/ultimos_bck.html', context)


@login_required
def listar_bd_view(request):
    """Listado de bases de datos usando BACKUPSGENERADOS"""
    try:
        # Consulta desde config
        resultados = ejecutar_consulta_personalizada(QUERIES['listar_bd_completo'])

        context = {
            'resultados': resultados,
        }

        return render(request, 'reportes/listar_bd.html', context)

    except Exception as e:
        logger.error(f"Error en listar BD: {e}")
        context = {
            'error': f'Error al cargar el listado de bases de datos: {str(e)}',
            'resultados': []
        }
        return render(request, 'reportes/listar_bd.html', context)


@login_required
@require_http_methods(["GET"])
def api_dashboard_metrics(request):
    """API para métricas del dashboard en tiempo real usando sp_DashboardMetrics"""
    try:
        # Usar el SP para obtener métricas
        datos_dashboard = ejecutar_sp_dashboard_metrics()
        metricas = datos_dashboard.get('metricas', {})
        
        return JsonResponse({
            'success': True,
            'data': {
                'total_servidores': metricas.get('total_servidores', 0),
                'total_bases_datos': metricas.get('total_bases_datos', 0),
                'backups_hoy': metricas.get('backups_hoy', 0),
                'backups_semana': metricas.get('backups_semana', 0),
                'total_backups_historico': metricas.get('total_backups_historico', 0),
                'total_jobs_historico': metricas.get('total_jobs_historico', 0),
            },
            'timestamp': timezone.now().isoformat()
        })

    except Exception as e:
        logger.error(f"Error en API metrics: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def disk_growth_view(request):
    """Reporte de crecimiento de discos usando DiskGrowthLog y sp_MonitorDiskGrowth"""
    try:
        # Filtros
        fecha_inicio = request.GET.get('fecha_inicio', (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
        fecha_fin = request.GET.get('fecha_fin', timezone.now().strftime('%Y-%m-%d'))
        servidor = request.GET.get('servidor', '')
        base_datos = request.GET.get('base_datos', '')
        
        logger.info(f"Ejecutando reporte de disk growth: {fecha_inicio} a {fecha_fin}")
        
        # Primero, ejecutar el SP para actualizar los datos más recientes
        try:
            ejecutar_procedimiento_almacenado('usp_MonitorDiskGrowth')
            logger.info("SP usp_MonitorDiskGrowth ejecutado exitosamente")
        except Exception as sp_error:
            logger.warning(f"Error ejecutando usp_MonitorDiskGrowth: {sp_error}")
            # Continuar con el reporte aunque falle el SP
        
        # Consultar datos históricos desde config
        params = [fecha_inicio, fecha_fin]
        query = QUERIES['disk_growth_detallado']
        
        # Aplicar filtros adicionales
        if servidor:
            query += " AND ServerIP LIKE %s"
            params.append(f'%{servidor}%')
        if base_datos:
            query += " AND DatabaseName LIKE %s"
            params.append(f'%{base_datos}%')
            
        query += " ORDER BY LogDate DESC, ServerIP, DatabaseName, FileName"
        
        resultados = ejecutar_consulta_personalizada(query, params)
        
        # Calcular estadísticas
        if resultados:
            total_registros = len(resultados)
            espacio_total_usado = sum(float(r.get('FileSizeMB', 0)) for r in resultados)
            espacio_total_libre = sum(float(r.get('DiskFreeMB', 0)) for r in resultados)
            servidores_unicos = len(set(r.get('ServerIP', '') for r in resultados))
            bases_unicas = len(set(r.get('DatabaseName', '') for r in resultados))
            
            # Identificar discos con poco espacio
            discos_criticos = len([r for r in resultados if r.get('status_class') == 'danger'])
            discos_advertencia = len([r for r in resultados if r.get('status_class') == 'warning'])
            
            estadisticas = {
                'total_registros': total_registros,
                'espacio_usado_gb': round(espacio_total_usado / 1024, 2),
                'espacio_libre_gb': round(espacio_total_libre / 1024, 2),
                'servidores': servidores_unicos,
                'bases_datos': bases_unicas,
                'discos_criticos': discos_criticos,
                'discos_advertencia': discos_advertencia
            }
        else:
            estadisticas = {
                'total_registros': 0,
                'espacio_usado_gb': 0,
                'espacio_libre_gb': 0,
                'servidores': 0,
                'bases_datos': 0,
                'discos_criticos': 0,
                'discos_advertencia': 0
            }
        
        # Paginación
        paginator = Paginator(resultados, PAGINATION['items_per_page'])
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        # Obtener lista de servidores únicos para el filtro
        try:
            servidores = ejecutar_consulta_personalizada(QUERIES['servidores_disk_growth'])
        except Exception:
            servidores = []

        # Obtener lista de bases de datos únicas para el filtro
        try:
            bases_datos = ejecutar_consulta_personalizada(QUERIES['bases_datos_disk_growth'])
        except Exception:
            bases_datos = []

        # Obtener tendencia de crecimiento (últimos 7 días)
        try:
            tendencia = ejecutar_consulta_personalizada(QUERIES['disk_growth_tendencia'])
            # Aplicar filtros de servidor y base de datos si están presentes
            if servidor or base_datos:
                tendencia = [
                    t for t in tendencia
                    if (not servidor or servidor.lower() in t.get('ServerIP', '').lower())
                    and (not base_datos or base_datos.lower() in t.get('DatabaseName', '').lower())
                ]
        except Exception:
            tendencia = []
        
        context = {
            'resultados': page_obj,
            'estadisticas': estadisticas,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'servidor': servidor,
            'base_datos': base_datos,
            'servidores': servidores,
            'bases_datos': bases_datos,
            'tendencia': tendencia[:50] if tendencia else []  # Limitar a 50 registros para el gráfico
        }
        
        return render(request, 'reportes/disk_growth.html', context)
        
    except Exception as e:
        logger.error(f"Error en disk growth: {e}")
        context = {
            'error': f'Error al cargar el reporte de crecimiento de discos: {str(e)}',
            'resultados': [],
            'estadisticas': {
                'total_registros': 0,
                'espacio_usado_gb': 0,
                'espacio_libre_gb': 0,
                'servidores': 0,
                'bases_datos': 0,
                'discos_criticos': 0,
                'discos_advertencia': 0
            },
            'fecha_inicio': fecha_inicio if 'fecha_inicio' in locals() else '',
            'fecha_fin': fecha_fin if 'fecha_fin' in locals() else '',
            'servidor': '',
            'base_datos': '',
            'servidores': [],
            'bases_datos': [],
            'tendencia': []
        }
        return render(request, 'reportes/disk_growth.html', context)


# =============================================================================
# EXPORTACIÓN A PDF
# =============================================================================

@login_required
def export_cumplimiento_pdf(request):
    """Exportar reporte de cumplimiento a PDF"""
    from .pdf_generator import generate_cumplimiento_pdf
    
    try:
        # Obtener parámetros
        fecha_inicio_param = request.GET.get('fecha')
        fecha_fin_param = request.GET.get('fecha1')
        
        # Configurar fechas
        if fecha_inicio_param:
            fecha_inicio = datetime.strptime(fecha_inicio_param, '%Y-%m-%d').strftime('%Y-%m-%d')
        else:
            fecha_inicio = datetime.now().replace(day=1).strftime('%Y-%m-%d')
            
        if fecha_fin_param:
            fecha_fin = datetime.strptime(fecha_fin_param, '%Y-%m-%d').strftime('%Y-%m-%d')
        else:
            fecha_fin = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
        
        # Obtener datos
        try:
            resultados = ejecutar_procedimiento_almacenado(
                'sp_Programaciondebcks',
                [fecha_inicio, fecha_fin]
            )

            # Normalizar resultados
            resultados = normalize_results(resultados, convert_cumplimiento_result)
        except Exception as e:
            logger.error(f"Error obteniendo datos para PDF: {e}")
            resultados = []
        
        # Calcular estadísticas
        total_registros = len(resultados)
        total_ejecutadas = sum(int(r.get('TOTAL', 0)) for r in resultados)
        total_programadas = sum(int(r.get('TOTALPROGRAM', 0)) for r in resultados)
        
        estadisticas = {
            'total_registros': total_registros,
            'total_ejecutadas': total_ejecutadas,
            'total_programadas': total_programadas,
            'promedio_cumplimiento': (total_ejecutadas / total_programadas * 100) if total_programadas > 0 else 0
        }
        
        # Generar PDF
        pdf_buffer = generate_cumplimiento_pdf(
            resultados, 
            estadisticas,
            fecha_inicio.replace('/', '-'),
            fecha_fin.replace('/', '-')
        )
        
        # Respuesta HTTP
        response = HttpResponse(pdf_buffer.read(), content_type='application/pdf')
        filename = f"cumplimiento_backup_{fecha_inicio.replace('/', '-')}_a_{fecha_fin.replace('/', '-')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        logger.info(f"PDF generado: {filename}")
        return response
        
    except Exception as e:
        logger.error(f"Error generando PDF de cumplimiento: {e}")
        messages.error(request, f'Error al generar el PDF: {str(e)}')
        return redirect('reportes:cumplimiento_backup')


@login_required
def export_jobs_pdf(request):
    """Exportar reporte de jobs a PDF"""
    from .pdf_generator import generate_jobs_pdf
    
    try:
        # Obtener parámetros
        fecha_inicio = request.GET.get('fecha_inicio', (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
        fecha_fin = request.GET.get('fecha_fin', timezone.now().strftime('%Y-%m-%d'))
        servidor = request.GET.get('servidor', '')
        resultado_filtro = request.GET.get('resultado', '')
        
        # Obtener datos
        try:
            resultados = ejecutar_procedimiento_almacenado(
                'sp_resultadoJobsBck',
                [fecha_inicio, fecha_fin]
            )

            # Normalizar resultados
            resultados = normalize_results(resultados, convert_jobs_result)
        except Exception as e:
            logger.error(f"Error obteniendo datos para PDF: {e}")
            resultados = []
        
        # Aplicar filtros
        if servidor:
            resultados = [r for r in resultados if servidor.lower() in r.get('SERVIDOR', '').lower()]
        if resultado_filtro:
            resultados = [r for r in resultados if resultado_filtro.lower() in r.get('RESULTADO', '').lower()]
        
        # Calcular estadísticas
        total = len(resultados)
        exitosos = len([r for r in resultados if 'exitoso' in r.get('RESULTADO', '').lower()])
        fallidos = len([r for r in resultados if 'fallido' in r.get('RESULTADO', '').lower() or 'error' in r.get('RESULTADO', '').lower()])
        
        stats = {
            'total': total,
            'exitosos': exitosos,
            'fallidos': fallidos,
            'porcentaje_exito': (exitosos / total * 100) if total > 0 else 0
        }
        
        # Generar PDF
        pdf_buffer = generate_jobs_pdf(resultados, stats, fecha_inicio, fecha_fin)
        
        # Respuesta HTTP
        response = HttpResponse(pdf_buffer.read(), content_type='application/pdf')
        filename = f"jobs_backup_{fecha_inicio}_a_{fecha_fin}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        logger.info(f"PDF generado: {filename}")
        return response
        
    except Exception as e:
        logger.error(f"Error generando PDF de jobs: {e}")
        messages.error(request, f'Error al generar el PDF: {str(e)}')
        return redirect('reportes:jobs_backup')


@login_required
def export_estados_pdf(request):
    """Exportar reporte de estados de BD a PDF"""
    from .pdf_generator import generate_estados_pdf
    
    try:
        # Obtener parámetros
        servidor = request.GET.get('servidor', '')
        estado = request.GET.get('estado', '')
        
        # Obtener datos
        try:
            # Ejecutar SP para actualizar datos
            try:
                ejecutar_procedimiento_almacenado('sp_MonitorDatabaseStatus')
            except Exception:
                pass
            
            # Consultar datos
            query = """
                SELECT 
                    ServerName as SERVIDOR,
                    DatabaseName as DATABASE_NAME,
                    LastLogDate as FECHA_DE_CREACION,
                    StateDesc as ESTADO,
                    [State] as TIPO_ESTADO,
                    ServerIP as IPSERVER
                FROM dbo.DatabaseStatusLog
                WHERE 1=1
            """
            params = []
            
            if servidor:
                query += " AND ServerName LIKE %s"
                params.append(f'%{servidor}%')
            if estado:
                query += " AND StateDesc LIKE %s"
                params.append(f'%{estado}%')
                
            query += " ORDER BY LastLogDate DESC, ServerName, DatabaseName"
            
            resultados = ejecutar_consulta_personalizada(query, params)
            
        except Exception as e:
            logger.error(f"Error obteniendo datos para PDF: {e}")
            resultados = []
        
        # Calcular estadísticas
        total = len(resultados)
        online = len([r for r in resultados if r.get('ESTADO', '').upper() == 'ONLINE'])
        servidores_unicos = len(set(r.get('SERVIDOR', '') for r in resultados if r.get('SERVIDOR')))
        
        stats = {
            'total': total,
            'online': online,
            'otros': total - online,
            'servidores': servidores_unicos
        }
        
        # Generar PDF
        pdf_buffer = generate_estados_pdf(resultados, stats)
        
        # Respuesta HTTP
        response = HttpResponse(pdf_buffer.read(), content_type='application/pdf')
        filename = f"estados_bd_{datetime.now().strftime('%Y-%m-%d')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        logger.info(f"PDF generado: {filename}")
        return response
        
    except Exception as e:
        logger.error(f"Error generando PDF de estados: {e}")
        messages.error(request, f'Error al generar el PDF: {str(e)}')
        return redirect('reportes:estados_db')


@login_required
def export_disk_growth_pdf(request):
    """Exportar reporte de crecimiento de discos a PDF"""
    from .pdf_generator import generate_disk_growth_pdf
    
    try:
        # Obtener parámetros
        fecha_inicio = request.GET.get('fecha_inicio', (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
        fecha_fin = request.GET.get('fecha_fin', timezone.now().strftime('%Y-%m-%d'))
        servidor = request.GET.get('servidor', '')
        base_datos = request.GET.get('base_datos', '')
        
        # Obtener datos
        try:
            query = """
                SELECT 
                    LogID,
                    ServerIP,
                    DatabaseName,
                    FileName,
                    FilePath,
                    FileSizeMB,
                    DiskFreeMB,
                    LogDate,
                    CAST((DiskFreeMB * 100.0 / (FileSizeMB + DiskFreeMB)) AS DECIMAL(5,2)) as PorcentajeLibre,
                    CASE 
                        WHEN DiskFreeMB < 10240 THEN 'danger'
                        WHEN DiskFreeMB < 51200 THEN 'warning'
                        ELSE 'success'
                    END as status_class
                FROM DiskGrowthLog
                WHERE CONVERT(date, LogDate) BETWEEN %s AND %s
            """
            
            params = [fecha_inicio, fecha_fin]
            
            if servidor:
                query += " AND ServerIP LIKE %s"
                params.append(f'%{servidor}%')
            if base_datos:
                query += " AND DatabaseName LIKE %s"
                params.append(f'%{base_datos}%')
                
            query += " ORDER BY LogDate DESC, ServerIP, DatabaseName, FileName"
            
            resultados = ejecutar_consulta_personalizada(query, params)
            
        except Exception as e:
            logger.error(f"Error obteniendo datos para PDF: {e}")
            resultados = []
        
        # Calcular estadísticas
        if resultados:
            total_registros = len(resultados)
            espacio_total_usado = sum(float(r.get('FileSizeMB', 0)) for r in resultados)
            espacio_total_libre = sum(float(r.get('DiskFreeMB', 0)) for r in resultados)
            discos_criticos = len([r for r in resultados if r.get('status_class') == 'danger'])
            discos_advertencia = len([r for r in resultados if r.get('status_class') == 'warning'])
            
            estadisticas = {
                'total_registros': total_registros,
                'espacio_usado_gb': round(espacio_total_usado / 1024, 2),
                'espacio_libre_gb': round(espacio_total_libre / 1024, 2),
                'discos_criticos': discos_criticos,
                'discos_advertencia': discos_advertencia
            }
        else:
            estadisticas = {
                'total_registros': 0,
                'espacio_usado_gb': 0,
                'espacio_libre_gb': 0,
                'discos_criticos': 0,
                'discos_advertencia': 0
            }
        
        # Generar PDF
        pdf_buffer = generate_disk_growth_pdf(resultados, estadisticas, fecha_inicio, fecha_fin)
        
        # Respuesta HTTP
        response = HttpResponse(pdf_buffer.read(), content_type='application/pdf')
        filename = f"disk_growth_{fecha_inicio}_a_{fecha_fin}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        logger.info(f"PDF generado: {filename}")
        return response
        
    except Exception as e:
        logger.error(f"Error generando PDF de disk growth: {e}")
        messages.error(request, f'Error al generar el PDF: {str(e)}')
        return redirect('reportes:disk_growth')


# =============================================================================
# EXPORTACIÓN A EXCEL CON FORMATO PROFESIONAL
# =============================================================================

@login_required
def export_cumplimiento_excel(request):
    """Exportar reporte de cumplimiento a Excel con formato profesional"""
    try:
        # Obtener parámetros
        fecha_inicio_param = request.GET.get('fecha')
        fecha_fin_param = request.GET.get('fecha1')
        
        # Configurar fechas
        if fecha_inicio_param:
            fecha_inicio = datetime.strptime(fecha_inicio_param, '%Y-%m-%d').strftime('%Y-%m-%d')
        else:
            fecha_inicio = datetime.now().replace(day=1).strftime('%Y-%m-%d')
            
        if fecha_fin_param:
            fecha_fin = datetime.strptime(fecha_fin_param, '%Y-%m-%d').strftime('%Y-%m-%d')
        else:
            fecha_fin = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
        
        # Obtener datos
        try:
            resultados = ejecutar_procedimiento_almacenado(
                'sp_Programaciondebcks',
                [fecha_inicio, fecha_fin]
            )

            # Normalizar y formatear resultados
            resultados = normalize_results(resultados, convert_cumplimiento_result)
            resultados = format_cumplimiento_results(resultados)
        except Exception as e:
            logger.error(f"Error obteniendo datos para Excel: {e}")
            resultados = []
        
        headers = ExportHeaders.CUMPLIMIENTO
        filename = ExportFileNames.timestamped('cumplimiento_backup', 'xlsx')
        title = f"{ReportTitles.CUMPLIMIENTO} ({fecha_inicio} - {fecha_fin})"

        logger.info(f"Excel generado: {filename}")
        return create_styled_excel(resultados, headers, filename, title=title, sheet_name=SheetNames.CUMPLIMIENTO)
        
    except Exception as e:
        logger.error(f"Error generando Excel de cumplimiento: {e}")
        messages.error(request, f'Error al generar el Excel: {str(e)}')
        return redirect('reportes:cumplimiento_backup')


@login_required
def export_jobs_excel(request):
    """Exportar reporte de jobs a Excel con formato profesional"""
    try:
        # Obtener parámetros
        fecha_inicio = request.GET.get('fecha_inicio', (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
        fecha_fin = request.GET.get('fecha_fin', timezone.now().strftime('%Y-%m-%d'))
        servidor = request.GET.get('servidor', '')
        resultado_filtro = request.GET.get('resultado', '')
        
        # Obtener datos
        try:
            resultados = ejecutar_procedimiento_almacenado(
                'sp_resultadoJobsBck',
                [fecha_inicio, fecha_fin]
            )

            # Normalizar resultados
            resultados = normalize_results(resultados, convert_jobs_result)
        except Exception as e:
            logger.error(f"Error obteniendo datos para Excel: {e}")
            resultados = []
        
        # Aplicar filtros
        if servidor:
            resultados = [r for r in resultados if servidor.lower() in r.get('SERVIDOR', '').lower()]
        if resultado_filtro:
            resultados = [r for r in resultados if resultado_filtro.lower() in r.get('RESULTADO', '').lower()]
        
        headers = ExportHeaders.JOBS
        filename = ExportFileNames.timestamped('jobs_backup', 'xlsx')
        title = f"{ReportTitles.JOBS} ({fecha_inicio} - {fecha_fin})"

        logger.info(f"Excel generado: {filename}")
        return create_styled_excel(resultados, headers, filename, title=title, sheet_name=SheetNames.JOBS)
        
    except Exception as e:
        logger.error(f"Error generando Excel de jobs: {e}")
        messages.error(request, f'Error al generar el Excel: {str(e)}')
        return redirect('reportes:jobs_backup')


@login_required
def export_estados_excel(request):
    """Exportar reporte de estados de BD a Excel con formato profesional"""
    try:
        # Obtener parámetros
        servidor = request.GET.get('servidor', '')
        estado = request.GET.get('estado', '')
        
        # Obtener datos
        try:
            # Ejecutar SP para actualizar datos
            try:
                ejecutar_procedimiento_almacenado('sp_MonitorDatabaseStatus')
            except Exception:
                pass
            
            # Consultar datos
            query = """
                SELECT 
                    ServerName as SERVIDOR,
                    DatabaseName as DATABASE_NAME,
                    LastLogDate as FECHA_DE_CREACION,
                    StateDesc as ESTADO,
                    [State] as TIPO_ESTADO,
                    ServerIP as IPSERVER
                FROM dbo.DatabaseStatusLog
                WHERE 1=1
            """
            params = []
            
            if servidor:
                query += " AND ServerName LIKE %s"
                params.append(f'%{servidor}%')
            if estado:
                query += " AND StateDesc LIKE %s"
                params.append(f'%{estado}%')
                
            query += " ORDER BY LastLogDate DESC, ServerName, DatabaseName"
            
            resultados = ejecutar_consulta_personalizada(query, params)
            
        except Exception as e:
            logger.error(f"Error obteniendo datos para Excel: {e}")
            resultados = []
        
        # Formatear fechas
        for r in resultados:
            if 'FECHA_DE_CREACION' in r and r['FECHA_DE_CREACION']:
                if hasattr(r['FECHA_DE_CREACION'], 'strftime'):
                    r['FECHA_DE_CREACION'] = r['FECHA_DE_CREACION'].strftime('%Y-%m-%d %H:%M:%S')
        
        headers = ExportHeaders.ESTADOS_DB
        filename = ExportFileNames.timestamped('estados_bd', 'xlsx')
        title = ReportTitles.ESTADOS_DB

        logger.info(f"Excel generado: {filename}")
        return create_styled_excel(resultados, headers, filename, title=title, sheet_name=SheetNames.ESTADOS_DB)
        
    except Exception as e:
        logger.error(f"Error generando Excel de estados: {e}")
        messages.error(request, f'Error al generar el Excel: {str(e)}')
        return redirect('reportes:estados_db')


@login_required
def export_disk_growth_excel(request):
    """Exportar reporte de crecimiento de discos a Excel con formato profesional"""
    try:
        # Obtener parámetros
        fecha_inicio = request.GET.get('fecha_inicio', (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
        fecha_fin = request.GET.get('fecha_fin', timezone.now().strftime('%Y-%m-%d'))
        servidor = request.GET.get('servidor', '')
        base_datos = request.GET.get('base_datos', '')
        
        # Obtener datos
        try:
            query = """
                SELECT 
                    LogID,
                    ServerIP,
                    DatabaseName,
                    FileName,
                    FilePath,
                    FileSizeMB,
                    DiskFreeMB,
                    LogDate,
                    CAST((DiskFreeMB * 100.0 / (FileSizeMB + DiskFreeMB)) AS DECIMAL(5,2)) as PorcentajeLibre,
                    CASE 
                        WHEN DiskFreeMB < 10240 THEN 'CRÍTICO'
                        WHEN DiskFreeMB < 51200 THEN 'ADVERTENCIA'
                        ELSE 'OK'
                    END as Estado
                FROM DiskGrowthLog
                WHERE CONVERT(date, LogDate) BETWEEN %s AND %s
            """
            
            params = [fecha_inicio, fecha_fin]
            
            if servidor:
                query += " AND ServerIP LIKE %s"
                params.append(f'%{servidor}%')
            if base_datos:
                query += " AND DatabaseName LIKE %s"
                params.append(f'%{base_datos}%')
                
            query += " ORDER BY LogDate DESC, ServerIP, DatabaseName, FileName"
            
            resultados = ejecutar_consulta_personalizada(query, params)
            
        except Exception as e:
            logger.error(f"Error obteniendo datos para Excel: {e}")
            resultados = []
        
        # Formatear datos
        for r in resultados:
            if 'LogDate' in r and r['LogDate']:
                if hasattr(r['LogDate'], 'strftime'):
                    r['LogDate'] = r['LogDate'].strftime('%Y-%m-%d %H:%M:%S')
            if 'FileSizeMB' in r and r['FileSizeMB'] is not None:
                r['FileSizeMB'] = f"{float(r['FileSizeMB']):,.2f}"
            if 'DiskFreeMB' in r and r['DiskFreeMB'] is not None:
                r['DiskFreeMB'] = f"{float(r['DiskFreeMB']):,.2f}"
            if 'PorcentajeLibre' in r and r['PorcentajeLibre'] is not None:
                r['PorcentajeLibre'] = f"{float(r['PorcentajeLibre']):.2f}%"
        
        headers = ExportHeaders.DISK_GROWTH
        filename = ExportFileNames.timestamped('crecimiento_discos', 'xlsx')
        title = f"{ReportTitles.DISK_GROWTH} ({fecha_inicio} - {fecha_fin})"

        logger.info(f"Excel generado: {filename}")
        return create_styled_excel(resultados, headers, filename, title=title, sheet_name=SheetNames.DISK_GROWTH)
        
    except Exception as e:
        logger.error(f"Error generando Excel de disk growth: {e}")
        messages.error(request, f'Error al generar el Excel: {str(e)}')
        return redirect('reportes:disk_growth')


# =============================================================================
# EXPORTACIÓN A CSV
# =============================================================================

@login_required
def export_cumplimiento_csv(request):
    """Exportar reporte de cumplimiento a CSV"""
    try:
        fecha_inicio_param = request.GET.get('fecha')
        fecha_fin_param = request.GET.get('fecha1')
        
        if fecha_inicio_param:
            fecha_inicio = datetime.strptime(fecha_inicio_param, '%Y-%m-%d').strftime('%Y-%m-%d')
        else:
            fecha_inicio = datetime.now().replace(day=1).strftime('%Y-%m-%d')
            
        if fecha_fin_param:
            fecha_fin = datetime.strptime(fecha_fin_param, '%Y-%m-%d').strftime('%Y-%m-%d')
        else:
            fecha_fin = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
        
        try:
            resultados = ejecutar_procedimiento_almacenado(
                'sp_Programaciondebcks',
                [fecha_inicio, fecha_fin]
            )

            # Normalizar y formatear resultados
            resultados = normalize_results(resultados, convert_cumplimiento_result)
            resultados = format_cumplimiento_results(resultados)
        except Exception as e:
            resultados = []
        
        headers = ExportHeaders.CUMPLIMIENTO
        filename = ExportFileNames.timestamped('cumplimiento_backup', 'csv')
        return create_csv_response(resultados, headers, filename)
        
    except Exception as e:
        logger.error(f"Error generando CSV: {e}")
        messages.error(request, f'Error al generar el CSV: {str(e)}')
        return redirect('reportes:cumplimiento_backup')


@login_required
def export_jobs_csv(request):
    """Exportar reporte de jobs a CSV"""
    try:
        fecha_inicio = request.GET.get('fecha_inicio', (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
        fecha_fin = request.GET.get('fecha_fin', timezone.now().strftime('%Y-%m-%d'))
        
        try:
            resultados = ejecutar_procedimiento_almacenado(
                'sp_resultadoJobsBck',
                [fecha_inicio, fecha_fin]
            )

            # Normalizar resultados
            resultados = normalize_results(resultados, convert_jobs_result)
        except Exception as e:
            resultados = []
        
        headers = ExportHeaders.JOBS
        filename = ExportFileNames.timestamped('jobs_backup', 'csv')
        return create_csv_response(resultados, headers, filename)
        
    except Exception as e:
        logger.error(f"Error generando CSV: {e}")
        messages.error(request, f'Error al generar el CSV: {str(e)}')
        return redirect('reportes:jobs_backup')


@login_required
def export_estados_csv(request):
    """Exportar reporte de estados a CSV"""
    try:
        servidor = request.GET.get('servidor', '')
        estado = request.GET.get('estado', '')
        
        try:
            query = """
                SELECT 
                    ServerName as SERVIDOR,
                    DatabaseName as DATABASE_NAME,
                    LastLogDate as FECHA_DE_CREACION,
                    StateDesc as ESTADO,
                    ServerIP as IPSERVER
                FROM dbo.DatabaseStatusLog
                WHERE 1=1
            """
            params = []
            
            if servidor:
                query += " AND ServerName LIKE %s"
                params.append(f'%{servidor}%')
            if estado:
                query += " AND StateDesc LIKE %s"
                params.append(f'%{estado}%')
                
            query += " ORDER BY LastLogDate DESC"
            
            resultados = ejecutar_consulta_personalizada(query, params)
        except Exception as e:
            resultados = []
        
        for r in resultados:
            if 'FECHA_DE_CREACION' in r and r['FECHA_DE_CREACION']:
                if hasattr(r['FECHA_DE_CREACION'], 'strftime'):
                    r['FECHA_DE_CREACION'] = r['FECHA_DE_CREACION'].strftime('%Y-%m-%d %H:%M:%S')
        
        # Usar solo los headers necesarios para CSV (sin TIPO_ESTADO para simplificar)
        headers = [h for h in ExportHeaders.ESTADOS_DB if h[0] != 'TIPO_ESTADO']
        filename = ExportFileNames.timestamped('estados_bd', 'csv')
        return create_csv_response(resultados, headers, filename)
        
    except Exception as e:
        logger.error(f"Error generando CSV: {e}")
        messages.error(request, f'Error al generar el CSV: {str(e)}')
        return redirect('reportes:estados_db')


@login_required
def export_disk_growth_csv(request):
    """Exportar reporte de crecimiento de discos a CSV"""
    try:
        fecha_inicio = request.GET.get('fecha_inicio', (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
        fecha_fin = request.GET.get('fecha_fin', timezone.now().strftime('%Y-%m-%d'))
        servidor = request.GET.get('servidor', '')
        base_datos = request.GET.get('base_datos', '')
        
        try:
            query = """
                SELECT 
                    ServerIP,
                    DatabaseName,
                    FileName,
                    FileSizeMB,
                    DiskFreeMB,
                    LogDate,
                    FilePath
                FROM DiskGrowthLog
                WHERE CONVERT(date, LogDate) BETWEEN %s AND %s
            """
            
            params = [fecha_inicio, fecha_fin]
            
            if servidor:
                query += " AND ServerIP LIKE %s"
                params.append(f'%{servidor}%')
            if base_datos:
                query += " AND DatabaseName LIKE %s"
                params.append(f'%{base_datos}%')
                
            query += " ORDER BY LogDate DESC"
            
            resultados = ejecutar_consulta_personalizada(query, params)
        except Exception as e:
            resultados = []
        
        for r in resultados:
            if 'LogDate' in r and r['LogDate']:
                if hasattr(r['LogDate'], 'strftime'):
                    r['LogDate'] = r['LogDate'].strftime('%Y-%m-%d %H:%M:%S')
        
        # Usar headers simplificados para CSV (sin PorcentajeLibre y Estado)
        headers = [h for h in ExportHeaders.DISK_GROWTH if h[0] not in ['PorcentajeLibre', 'Estado']]
        filename = ExportFileNames.timestamped('crecimiento_discos', 'csv')
        return create_csv_response(resultados, headers, filename)
        
    except Exception as e:
        logger.error(f"Error generando CSV: {e}")
        messages.error(request, f'Error al generar el CSV: {str(e)}')
        return redirect('reportes:disk_growth')
